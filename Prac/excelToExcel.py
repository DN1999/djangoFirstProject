from openpyxl import Workbook, load_workbook

# Create a new workbook and select the active worksheet
wb1 = Workbook()
ws1 = wb1.active

# Write data to the first workbook
ws1['A1'] = 'Name'
ws1['B1'] = 'Age'
ws1['A2'] = 'John'
ws1['B2'] = 25
ws1['A3'] = 'Alice'
ws1['B3'] = 30

# Save the first workbook
wb1.save('source.xlsx')

# Load the source workbook
wb2 = load_workbook('source.xlsx')
ws2 = wb2.active

# Create a new workbook to write data from the source workbook
wb3 = Workbook()
ws3 = wb3.active

# Copy data from the source workbook to the new workbook
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
    new_row = []
    for cell in row:
        new_row.append(cell.value)
    ws3.append(new_row)

# Save the new workbook
wb3.save('destination.xlsx')

# Read data from the new workbook
wb4 = load_workbook('destination.xlsx')
ws4 = wb4.active

# Display data from the new workbook
print("Data from destination.xlsx:")
for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=ws4.max_column):
    for cell in row:
        print(cell.value, end="\t")
    print()

# Close workbooks
wb2.close()
wb4.close()
