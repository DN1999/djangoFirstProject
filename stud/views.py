from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import workbook,load_workbook
import pandas as pd

# Create your views here.
def home(request):
    # return HttpResponse("<h1>Krishna World</h1>");
  
    # return render(request,"stud/home.html",{"name":"सर्वांना"})
    return render(request, 'stud/home.html', {})

def student(request):
      # Read Excel file
    excel_data = None
    try:
        # Assuming the Excel file is named studentRecord.xlsx and is located in the stud folder
        excel_data = pd.read_excel('./stud/studentRecord.xlsx')
        # Convert data to HTML table
        excel_html = excel_data.to_html(classes='table table-striped', index=False)
    except Exception as e:
        error_message = f"An error occurred: {e}"
        return render(request, 'stud/home.html', {'error_message': error_message})

    return render(request,"stud/student.html",{'excel_html': excel_html})

def addStudentPage(request):
    res=""
    return render(request,"stud/add_student.html",{"result":res})

def register(request):
    name= request.POST["name"]
    email= request.POST["email"]
    gender= request.POST["gender"]

#   Create a new workbook and select the active worksheet
    # wb1 = Workbook()
    wb1 = load_workbook('./stud/studentRecord.xlsx')
    ws1 = wb1.active
    
    # ws1['A1'] = 'Name'
    # ws1['B1'] = 'Email'
    # ws1['C1'] = 'Gender'
    last=ws1.max_row +1
    ws1[f'A{last}']=ws1.max_row
    ws1[f'B{last}']=name
    ws1[f'C{last}']=email
    ws1[f'D{last}']=gender

    wb1.save('./stud/studentRecord.xlsx')

    res = "Student Registered Successfully....!"
    return render(request,"stud/add_student.html",{"result":res})
#End Stuent Register Function

#Search Student data
def searchStudent(request):
    wb1=load_workbook('./stud/studentRecord.xlsx')
    ws1=wb1.active
    srno=request.POST["srno"]    
    index=int(srno)+1
    # srno=ws1[f'A{index}']
    name=ws1[f'B{index}'].value
    email=ws1[f'C{index}'].value
    gender=ws1[f'D{index}'].value    

    wb1.close
    # name="Dhananjay"
    # email="dhananjaypnaphade@gamil.com"
    # gender="female"
    return render(request,"stud/update_student.html",{"res_srno":srno,"res_name":name,"res_email":email,"res_gender":gender})

# Upadate Student Page data Render
def updateStudentPage(request):
    return render(request,"stud/update_student.html",{})

#start Student Update Data Function
def updateStudent(request):
    wb1=load_workbook('./stud/studentRecord.xlsx')
    ws1=wb1.active
    srno=request.POST["srno"]
    name=request.POST["name"]
    email=request.POST['email']
    gender= request.POST['gender']
    index=int(srno)+1
    ws1[f'A{index}']=int(srno)
    ws1[f'B{index}']=name
    ws1[f'C{index}']=email
    ws1[f'D{index}']=gender

    wb1.save('./stud/studentRecord.xlsx')
    res="Student Data Updated Succesfully"
    return render(request,"stud/update_student.html",{"result":res})

